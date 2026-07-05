"""Phase 3 — Excel generation with openpyxl.

Builds a workbook with a Summary sheet followed by one sheet per element type.
Formatting: bold header, frozen header row, auto-ish column widths.
Values are written verbatim (faithful 1:1).
"""
from __future__ import annotations

from datetime import datetime
from typing import List

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

from .extractor import ExtractionResult

_HEADER_FONT = Font(bold=True, color="FFFFFF")
_HEADER_FILL = PatternFill("solid", fgColor="305496")
_MAX_WIDTH = 70
_MIN_WIDTH = 10


def _autofit(ws, columns: List[str], rows) -> None:
    for idx, col in enumerate(columns, start=1):
        longest = len(str(col))
        for r in rows:
            v = r.get(col, "")
            if v:
                longest = max(longest, len(str(v)))
        width = max(_MIN_WIDTH, min(_MAX_WIDTH, longest + 2))
        ws.column_dimensions[get_column_letter(idx)].width = width


def _style_header(ws, ncols: int) -> None:
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(vertical="center")
    ws.freeze_panes = "A2"


def _write_data_sheet(wb: Workbook, title: str, columns: List[str], rows) -> None:
    ws = wb.create_sheet(title=title[:31])  # Excel sheet-name limit
    ws.append(columns)
    for r in rows:
        ws.append([r.get(col, "") for col in columns])
    _style_header(ws, len(columns))
    _autofit(ws, columns, rows)


def _write_summary(wb: Workbook, result: ExtractionResult) -> None:
    ws = wb.active
    ws.title = "Summary"
    ws.append(["FP2Any — Forcepoint Configuration Extract"])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([])
    ws.append(["Source XML file", result.source_filename])
    ws.append(["Export build", result.meta.get("build", "")])
    ws.append(["Update package version", result.meta.get("update_package_version", "")])
    ws.append(["Conversion timestamp", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
    ws.append(["Total top-level elements", result.total_elements])
    ws.append([])

    hdr_row = ws.max_row + 1
    ws.append(["Sheet", "Element count"])
    for c in range(1, 3):
        cell = ws.cell(row=hdr_row, column=c)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
    for sheet, count in result.counts.items():
        ws.append([sheet, count])

    if result.unknown_tags:
        ws.append([])
        ws.append(["Unknown/unsupported tags (routed to Other_Elements)"])
        ws.cell(row=ws.max_row, column=1).font = Font(bold=True, italic=True)
        ws.append(["Tag", "Count"])
        for tag, count in sorted(result.unknown_tags.items()):
            ws.append([tag, count])

    ws.column_dimensions["A"].width = 45
    ws.column_dimensions["B"].width = 20


def build_workbook(result: ExtractionResult, progress=None) -> Workbook:
    """Build the full workbook. ``progress`` is an optional callback
    ``(pct: int, stage: str)`` reporting each sheet as it is written."""
    wb = Workbook()
    _write_summary(wb, result)
    total = len(result.sheets) or 1
    for i, (sheet, (columns, rows)) in enumerate(result.sheets.items()):
        if progress:
            progress(int(i * 100 / total), sheet)
        _write_data_sheet(wb, sheet, columns, rows)
    if progress:
        progress(100, "Done")
    return wb


def build_selected_workbook(result: ExtractionResult, sheet_names) -> Workbook:
    """Workbook containing only the given data sheet(s), no Summary.

    Used by the web UI to download individual sheets.
    """
    wb = Workbook()
    default = wb.active
    for sheet in sheet_names:
        if sheet in result.sheets:
            columns, rows = result.sheets[sheet]
            _write_data_sheet(wb, sheet, columns, rows)
    # drop the empty default sheet if we added at least one real sheet
    if len(wb.sheetnames) > 1 and default.title == "Sheet":
        wb.remove(default)
    return wb


def write_excel(result: ExtractionResult, output_path: str) -> str:
    wb = build_workbook(result)
    wb.save(output_path)
    return output_path
